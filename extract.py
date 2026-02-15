import openpyxl, json

wb = openpyxl.load_workbook('Uren_per_Bedrijf_ITFin_V5.xlsx', data_only=True)
result = {}

for sheet_name in wb.sheetnames:
    if sheet_name == 'Overzicht':
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    time_entries = []
    tickets = []
    section = None

    for row in rows:
        vals = list(row)
        # Detect section
        if vals[0] == 'Uren registraties':
            section = 'hours_header'
            continue
        if vals[0] == 'Zoho Desk tickets':
            section = 'tickets_header'
            continue
        if vals[0] == 'Geen uren registraties' or vals[0] == 'Geen tickets':
            continue

        # Hours section
        if section == 'hours_header' and vals[0] == 'Medewerker':
            section = 'hours'
            continue
        if section == 'hours':
            if vals[0] is None or vals[2] == 'Totaal:':
                continue
            time_entries.append({
                'employee': str(vals[0]) if vals[0] else None,
                'date': str(vals[1]) if vals[1] else None,
                'company': str(vals[2]) if vals[2] else None,
                'hours': float(vals[3]) if vals[3] else 0,
                'description': str(vals[4]) if vals[4] else None,
            })

        # Tickets section
        if section == 'tickets_header' and vals[0] == 'Ticket #':
            section = 'tickets'
            continue
        if section == 'tickets' and vals[0] is not None:
            tickets.append({
                'ticket_nr': int(vals[0]) if vals[0] else None,
                'date': str(vals[1]) if vals[1] else None,
                'subject': str(vals[2]) if vals[2] else None,
                'tasks_performed': str(vals[3]) if vals[3] else None,
                'pc_name': str(vals[4]) if vals[4] else None,
                'serial_number': str(vals[5]) if vals[5] else None,
                'pending_tasks': str(vals[6]) if vals[6] else None,
                'equipment': str(vals[7]) if vals[7] else None,
                'time_hours': float(vals[8]) if vals[8] else None,
                'ticket_date': str(vals[9]) if vals[9] else None,
                'status': str(vals[10]) if vals[10] else None,
            })

    result[sheet_name] = {
        'time_entries': time_entries,
        'tickets': tickets
    }

with open('extracted_data.json', 'w') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

# Print summary
total_hours = 0
total_tickets = 0
for company, data in result.items():
    te = len(data['time_entries'])
    tk = len(data['tickets'])
    total_hours += te
    total_tickets += tk
    print(f"{company}: {te} uren, {tk} tickets")

print(f"\nTOTAAL: {total_hours} uren registraties, {total_tickets} tickets")
